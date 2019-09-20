import sys, getopt
import codecs
import os
import openpyxl
from openpyxl import load_workbook

sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

def marge(inputfile,inputfile1):
    #wb = load_workbook('/Users/benyang/Downloads/GT/ticket.xlsx')
    wb = load_workbook(inputfile)
    sheet_names = wb.sheetnames
    ws=wb[sheet_names[0]]

    str1=""
    str2=""
    '''
    s=ws['A2']
    #print (s.value)
    for letter in range (0,len(s.value)-2):
        if letter<2:
            str1+=s.value[letter]

        else:
            str2+=s.value[letter]

    print (str1," ",str2)

    '''
    #wb1 = load_workbook('/Users/benyang/Downloads/GT/form.xlsx')
    wb1 = load_workbook(inputfile1)
    #wb1.template = True
    #wb2 = Workbook()
    #print(wb.get_sheet_names())
    #ws=wb['ticket']
    #ticket = wb.active

    sheet_names1 = wb1.sheetnames
    ws1=wb1[sheet_names1[0]]
    #ws1=wb['1_1']
    #ws1=wb[sheet_names[1]]
    #print (ws['C2'].value)
    #os.system('pause')
    i=1
    j=0
    k=0

    for num in range(ws.min_row+1,ws.max_row):
    #for num in range(ws.min_row+1,52):

        t=ws['B'+str(num)]
        t1=ws['C'+str(num)]
        t2=ws['D'+str(num)]
        t3=ws['E'+str(num)]



        if j==0:
            #print (t1.value)
            #if isinstance(t.value,str):
                #ws1['B'+str(i+11)]=t.value.encode('UTF-8')
                #print (t.value.encode('UTF-8'))
            #if isinstance(t.value,int):
            ws1['B'+str(i+11)]=t.value
                #print (t.value)
            #if str(t1.value)=='V':
            ws1['K'+str(i+11)]=t1.value
            ws1['G'+str(i+11)]=t2.value
            ws1['N'+str(i+11)]=t3.value
            if i==25:
                i=1
                j=1
            else:
                i+=1
        else:
            #print (t1.value)
            #if isinstance(t.value,str):
                #ws1['S'+str(i+11)]=t.value.encode('UTF-8')
                #print (t.value.encode('UTF-8'))
            #if isinstance(t.value,int):
            ws1['S'+str(i+11)]=t.value
                #print (t.value)
            #if str(t1.value)=='V':
            ws1['X'+str(i+11)]=t1.value
            #else:
                #ws1['X'+str(i+11)]=t1.value
            ws1['V'+str(i+11)]=t2.value
            ws1['AA'+str(i+11)]=t3.value
            if i==25:
                i=1
                j=0
                k+=1
                s=ws['A'+str(num)]
                for letter in range (0,len(s.value)-2):
                    if letter<2:
                        str1+=s.value[letter]

                    else:
                        str2+=s.value[letter]
                ws1['X8']=str1
                ws1['AA8']=str2[0]
                ws1['AB8']=str2[1]
                ws1['AC8']=str2[2]
                ws1['AD8']=str2[3]
                ws1['AE8']=str2[4]
                ws1['AF8']=str2[5]
                str1=""
                str2=""
                ws1=wb1[sheet_names1[k]]
            else:
                i+=1

    wb1.save(r'/Users/benyang/Downloads/GT/form_new.xlsx')

def main(argv):
   inputfile = ''
   inputfile1 = ''
   outputfile = ''
   try:
      opts, args = getopt.getopt(argv,"hi:j:o:",["ifile=","ifile1=","ofile="])
   except getopt.GetoptError:
      print ('test.py -i <inputfile> -o <inputfile1>')
      sys.exit(2)
   for opt, arg in opts:
      if opt == '-h':
         print ('test.py -i <inputfile> -o <inputfile1>')
         sys.exit()
      elif opt in ("-i", "--ifile"):
         inputfile = arg
      elif opt in ("-j", "--ifile1"):
         inputfile1 = arg
      elif opt in ("-o", "--ofile"):
         outputfile = arg
   print (inputfile,"-",inputfile1)

   if (inputfile!='' and inputfile1!=''):
       marge(inputfile,inputfile1)

if __name__ == "__main__":
   main(sys.argv[1:])
   print ('Done!!')
#for row in ws.iter_rows(min_row=ws.min_row+1,max_row=ws.max_row,values_only=True):
#for row in ws.iter_rows(min_row=1, max_col=3, max_row=2,values_only=True):
#for row in ws.values:

    #for value in row:
        #print (value)
        #if isinstance(value,str):
            #print (value.encode('UTF-8'))
        #if isinstance(value,int):
            #print (value)

#t=ws['B108']
#print (isinstance(t.value,int))
#print (type(t))
#print (t.value.encode('UTF-8'))
#for c in ws['B1:B3']:
    #print (c)
#column=ticket.columns
#print (ws.min_row,ws.max_row)
#for row in ws.iter_rows('B{}:B{}'.format(ws.min_row+1,ws.max_row)):
    #for cell in row:
        #print (cell.value)

#for cell in column:
    #print (cell.value)
#for row in ws.iter_rows(min_row=2,max_row=3):

    #print (row)
    #for cell in row:
     #if row == 'B':
     #print (cell)
      #print (cell.value)
