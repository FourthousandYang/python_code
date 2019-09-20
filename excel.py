import sys
import codecs
import os
import openpyxl
from openpyxl import load_workbook

sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

wb = load_workbook('/Users/benyang/Downloads/GT/ticket1.xlsx')
sheet_names = wb.sheetnames
ws=wb[sheet_names[0]]

str1=""
str2=""
sum=0

year=input('year:')
months=input('months:')
month=input('month:')
day=input('day:')
'''

t1=ws['I'+str(41)]
print (len(str(t1.value)))
if str(t1.value).count==8:
    print ('1234')
else:
    print (t1.value)



'''
wb1 = load_workbook('/Users/benyang/Downloads/GT/form.xlsx')
#wb1.template = True
#wb2 = Workbook()
#print(wb.get_sheet_names())
#ws=wb['ticket']
#ticket = wb.active

sheet_names1 = wb1.sheetnames
ws1=wb1[sheet_names1[0]]
#ws1=wb['1_1']
#ws1=wb[sheet_names[1]]
#print (ws['C2'].value)
#os.system('pause')


i=1
j=0
k=0
l=1
m=0
end=ws.max_row-1
for num in range(ws.min_row+1,3002):
#for num in range(ws.min_row+1,52):
    ws1=wb1[sheet_names1[k]]
    t=ws['I'+str(num)]
    t1=ws['J'+str(num)]
    t2=ws['K'+str(num)]
    t3=ws['L'+str(num)]
    t4=ws['N'+str(num)]



    if j==0:
        #print (t1.value)
        #if isinstance(t.value,str):
            #ws1['B'+str(i+11)]=t.value.encode('UTF-8')
            #print (t.value.encode('UTF-8'))
        #if isinstance(t.value,int):
        if num<=end:
            if t.value=='________':
                ws1['B'+str(i+11)]='********'
                ws1['N'+str(i+11)]='-'
                sum+=t2.value+t3.value
                ws1['G'+str(i+11)]=t2.value+t3.value
            elif t.value=='作廢':
                ws1['B'+str(i+11)]=str(t.value)
            else:
                if len(str(t.value))==8:
                    ws1['B'+str(i+11)]=str(t.value)
                else:
                    ws1['B'+str(i+11)]='0'+str(t.value)
                ws1['N'+str(i+11)]=t3.value
                ws1['G'+str(i+11)]=t2.value
                #print (t.value)
            #if str(t1.value)=='V':
            ws1['K'+str(i+11)]=t1.value
            ws1['P'+str(i+11)]=t4.value
        else:
            if num==end+1:
                ws1['B'+str(i+11)]="以下空白"
            else:
                ws1['B'+str(12)]="以下空白"
            ws1['N'+str(i+11)]='-'
            ws1['G'+str(i+11)]='-'

        if i==25:
            i=1
            j=1
        else:
            i+=1
    else:
        #print (t1.value)
        #if isinstance(t.value,str):
            #ws1['S'+str(i+11)]=t.value.encode('UTF-8')
            #print (t.value.encode('UTF-8'))
        #if isinstance(t.value,int):
        if num<=end:
            if t.value=='________':
                ws1['S'+str(i+11)]='********'
                ws1['AA'+str(i+11)]='-'
                sum+=t2.value+t3.value
                ws1['V'+str(i+11)]=t2.value+t3.value
            elif t.value=='作廢':
                ws1['S'+str(i+11)]=str(t.value)

            else:
                if len(str(t.value))==8:
                    ws1['S'+str(i+11)]=str(t.value)
                else:
                    ws1['S'+str(i+11)]='0'+str(t.value)
                ws1['AA'+str(i+11)]=t3.value
                ws1['V'+str(i+11)]=t2.value
        #ws1['S'+str(i+11)]=t.value
            #print (t.value)
        #if str(t1.value)=='V':
            ws1['X'+str(i+11)]=t1.value
            ws1['AD'+str(i+11)]=t4.value
        else:
            if num==end+1:
                ws1['S'+str(i+11)]="以下空白"

            ws1['AA'+str(i+11)]='-'
            ws1['V'+str(i+11)]='-'
        #else:
            #ws1['X'+str(i+11)]=t1.value

        #ws1['AA'+str(i+11)]=t3.value
        if i==25:
            i=1
            j=0
            k+=1

            if num<=end:
                s=ws['H'+str(num)]
                for letter in range (0,len(s.value)-2):
                    if letter<2:
                        str1+=s.value[letter]

                    else:
                        str2+=s.value[letter]
                ws1['X8']=str1
                ws1['AA8']=str2[0]
                ws1['AB8']=str2[1]
                ws1['AC8']=str2[2]
                ws1['AD8']=str2[3]
                ws1['AE8']=str2[4]
                ws1['AF8']=str2[5]
                ws1['Z41']=sum
            else:

                s=ws['H'+str(end)]
                #ss=str(int(s.value)+l)
                for letter in range (0,len(s.value)-2):
                    if letter<2:
                        str1+=s.value[letter]

                    else:
                        str2+=s.value[letter]
                ws1['X8']=str1
                str2=str(int(str2)+l)
                ws1['AA8']=str2[0]
                ws1['AB8']=str2[1]
                ws1['AC8']=str2[2]
                ws1['AD8']=str2[3]
                ws1['AE8']=str2[4]
                ws1['AF8']=str2[5]
                if m==0:
                    m=1
                else:
                    l+=1
                    m=0
            str1=""
            str2=""
            sum=0
            eq="=COUNTIF(B12:F36,"+'"作廢"'+")+COUNTIF(S12:U36,"+'"作廢"'+")"
            eq1="=COUNTIF(B12:B36,"+'""'+")+COUNTIF(S12:S36,"+'""'+")"+"+COUNTIF(B12:B36,"+'"以下空白"'+")+COUNTIF(S12:S36,"+'"以下空白"'+")"
            ws1['T41']=eq#"=COUNTIF(B12:F36,"+"")+COUNTIF(S12:U36,)"
            ws1['U41']=eq1
            ws1['J45']="本表為本期(月)電子計算機發票明細表共 "+"60"+" 張之第         "+str(k)
            ws1['A3']="民國    "+str(year)+"     年    "+str(months)+"    月份"
            ws1['B43']="申報日期:    "+str(year)+"    年    "+str(month)+"   月    "+str(day)+"    日"
            #if k==0:

            #else:
            ws1['P42']=""
            ws1['Q42']=""
            ws1['S42']=""
            ws1['T42']=""
            ws1['U42']=""
            ws1['V42']=""
            ws1['W42']=""
            ws1['Z42']=""
            #if num!=3002:
                #ws1=wb1[sheet_names1[k]]
        else:
            i+=1
ws1=wb1[sheet_names1[0]]
ws1['P42']="=SUM('1_1:30_2'!P41)"
ws1['Q42']="=SUM('1_1:30_2'!Q41)"
ws1['S42']="=SUM('1_1:30_2'!S41)"
ws1['T42']="=SUM('1_1:30_2'!T41)"
ws1['U42']="=SUM('1_1:30_2'!U41)"
ws1['V42']="=SUM('1_1:30_2'!V41)"
ws1['W42']="=SUM('1_1:30_2'!W41)"
ws1['Z42']="=SUM('1_1:30_2'!Z41)"
wb1.save(r'/Users/benyang/Downloads/GT/form_new.xlsx')

#for row in ws.iter_rows(min_row=ws.min_row+1,max_row=ws.max_row,values_only=True):
#for row in ws.iter_rows(min_row=1, max_col=3, max_row=2,values_only=True):
#for row in ws.values:

    #for value in row:
        #print (value)
        #if isinstance(value,str):
            #print (value.encode('UTF-8'))
        #if isinstance(value,int):
            #print (value)

#t=ws['B108']
#print (isinstance(t.value,int))
#print (type(t))
#print (t.value.encode('UTF-8'))
#for c in ws['B1:B3']:
    #print (c)
#column=ticket.columns
#print (ws.min_row,ws.max_row)
#for row in ws.iter_rows('B{}:B{}'.format(ws.min_row+1,ws.max_row)):
    #for cell in row:
        #print (cell.value)

#for cell in column:
    #print (cell.value)
#for row in ws.iter_rows(min_row=2,max_row=3):

    #print (row)
    #for cell in row:
     #if row == 'B':
     #print (cell)
      #print (cell.value)
